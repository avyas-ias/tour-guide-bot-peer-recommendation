from openpyxl import load_workbook
import psycopg2

conn = psycopg2.connect(
   database="dvk6lrjqokq93", 
   user='zubvkjwkjfqnni', 
   password='efa884e7ff57cf3f2c11301bea345029c62890dd72cac8c743057a4a910184f0', 
   host='ec2-44-194-112-166.compute-1.amazonaws.com', 
   port= '5432'
)
cursor = conn.cursor()

workbook = load_workbook(filename="./integralad.xlsx")
sheet = workbook.active
count = 0
for row in sheet.iter_rows(values_only=True,min_row=2):

    count += 1
    lname, fname = row[0].split(',')
    lname = lname.strip().lower()
    fname = fname.strip().lower()
    name = f"{fname} {lname}"

    try:
        reports_to_lname,reports_to_fname = row[1].split(',')
        reports_to_fname = reports_to_fname.strip().lower()
        reports_to_lname = reports_to_lname.strip().lower()
        reports_to = f"{reports_to_fname} {reports_to_lname}"
    except:
        reports_to_fname = "0"
        reports_to_lname = "0"
        reports_to = f"{reports_to_fname}"


    job_title = row[2].lower()

    number_of_direct = int(row[3])

    dept = row[4].lower()

    location = row[5].lower()

    cursor.execute(f"insert into org_chart values(default,'{name}','{reports_to}','{job_title}',{number_of_direct},'{dept}','{location}');")
    print(f"[{count}] inserted values(default,{name},{reports_to},{job_title},{number_of_direct},{dept},{location});")

print("Commiting the DB changes !")
conn.commit()